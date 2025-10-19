import os
import pickle
import pandas as pd
import torch
import sys
from torch import nn
import openpyxl

sys.path.append('/path/to/root')

from models.TwoStage1D import Predictor
from utils.multi_stage_train import setup_seed, rm_module
from datasets.dataprocess import getseq_input_single

# set device
os.environ['CUDA_VISIBLE_DEVICES'] = '0,1'

# set random seed
setup_seed(seed=666)

# set parameters
dropout_rate = 0.2
predictor_model_path = './root/to/ckpt_path'
datapath = './root/to/data_path'
# prepare scaler and dataset
with open('./path/to/temp_data', 'rb') as temp:
    temp_dict = pickle.load(temp)
scaler = {'mean': temp_dict['mean'], 'std': temp_dict['std']}

# get trained model
predictor_trained = Predictor(dropout_rate=dropout_rate)
predictor_state_dict_rm = rm_module(predictor_model_path)
predictor_trained.load_state_dict(predictor_state_dict_rm, strict=True)

predictor_trained = nn.DataParallel(predictor_trained).cuda().eval()

df = pd.read_excel(datapath).iloc[:, :2]
df = df.dropna()
# seq_id = df['name'].tolist()
seqs = df['seqs'].tolist()
specific_labels = df['specific_labels'].tolist()

rl_list = []
for i in range(len(seqs)):
    seq_input = getseq_input_single(seqs[i][:100], max_len=100)
    seq_input = seq_input.cuda()
    with torch.no_grad():
        rl = predictor_trained(seq_input)
        rl = rl.item()
        rl = rl * scaler['std'] + scaler['mean']
        rl_list.append(rl)

# save results
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.cell(row=1, column=1).value = 'name'
worksheet.cell(row=1, column=2).value = 'seq'
worksheet.cell(row=1, column=3).value = 'specific_label'
worksheet.cell(row=1, column=4).value = 'rl_pred'
for i in range(len(seqs)):
    worksheet.cell(row=i+2, column=1).value = i+1
    worksheet.cell(row=i+2, column=2).value = seqs[i]
    worksheet.cell(row=i+2, column=3).value = specific_labels[i]
    worksheet.cell(row=i+2, column=4).value = rl_list[i]
workbook.save('./path/to/save_path')